from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from models import db
from models.project import Project
from models.application import Application
from models.checkin import Checkin
from models.user import User


def _parse_date(s):
    """与 statistics_service 同款日期解析；解析失败返回 None。"""
    if not s:
        return None
    if isinstance(s, datetime):
        return s
    try:
        return datetime.strptime(str(s)[:10], '%Y-%m-%d')
    except (ValueError, TypeError):
        return None


def _apply_project_filters(query, start_date=None, end_date=None,
                           category=None, college_id=None):
    """对项目相关查询追加日期 / 类型 / 学院筛选。

    与 statistics_service 的筛选语义保持一致：
    - 日期范围按 Project.start_time
    - 学院通过该项目存在 approved 报名人匹配
    """
    sd = _parse_date(start_date)
    ed = _parse_date(end_date)
    if sd:
        query = query.filter(Project.start_time >= sd)
    if ed:
        query = query.filter(Project.start_time <= ed)
    if category:
        query = query.filter(Project.category == category)
    if college_id:
        sub = db.session.query(Application.project_id).join(
            User, User.id == Application.user_id
        ).filter(
            Application.status == 'approved',
            User.college_id == college_id,
        ).distinct().subquery()
        query = query.filter(Project.id.in_(sub.select()))
    return query


def export_application_list(project_id):
    """导出某项目报名名单"""
    project = Project.query.get(project_id)
    if not project:
        raise ValueError('项目不存在')

    apps = Application.query.filter_by(project_id=project_id).filter(
        Application.status != 'cancelled'
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = '报名名单'
    ws.append(['序号', '姓名', '学号', '学院', '专业', '手机号', '报名说明', '状态', '报名时间'])

    for i, a in enumerate(apps, 1):
        ws.append([
            i, a.user.real_name, a.user.student_id, a.user.college_name,
            a.user.major, a.user.phone, a.apply_reason or '',
            a.STATUS_LABELS.get(a.status, ''),
            a.apply_time.strftime('%Y-%m-%d %H:%M') if a.apply_time else '',
        ])

    # 调整列宽
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 15

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, f'{project.title}_报名名单.xlsx'


def export_project_stats(start_date=None, end_date=None, category=None, college_id=None):
    """导出项目统计数据（支持按时间区间 / 类型 / 学院筛选，与统计页保持一致）"""
    query = Project.query.filter_by(is_deleted=False)
    query = _apply_project_filters(query, start_date, end_date, category, college_id)
    projects = query.order_by(Project.start_time.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = '项目统计'
    ws.append(['序号', '项目名称', '类型', '状态', '招募人数', '已通过人数', '签到人数', '创建人', '创建时间'])

    for i, p in enumerate(projects, 1):
        approved = Application.query.filter_by(project_id=p.id, status='approved').count()
        checked = Checkin.query.filter_by(project_id=p.id).filter(Checkin.sign_out_time.isnot(None)).count()
        ws.append([
            i, p.title, p.category or '', p.STATUS_LABELS.get(p.status, ''),
            p.max_people, approved, checked,
            p.creator.real_name if p.creator else '',
            p.created_at.strftime('%Y-%m-%d') if p.created_at else '',
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 15

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, '项目统计数据.xlsx'


def export_hours_records(start_date=None, end_date=None, category=None, college_id=None):
    """导出志愿者时长记录（支持按时间区间 / 类型 / 学院筛选）

    日期范围按 Checkin.sign_in_time 过滤（与统计页月度趋势同口径）。
    """
    query = Checkin.query.filter_by(status='confirmed')

    sd = _parse_date(start_date)
    ed = _parse_date(end_date)
    if sd:
        query = query.filter(Checkin.sign_in_time >= sd)
    if ed:
        query = query.filter(Checkin.sign_in_time <= ed)
    if category:
        query = query.join(Project, Checkin.project_id == Project.id).filter(
            Project.category == category
        )
    if college_id:
        query = query.join(User, User.id == Checkin.user_id).filter(
            User.college_id == college_id
        )

    checkins = query.all()

    wb = Workbook()
    ws = wb.active
    ws.title = '时长记录'
    ws.append(['序号', '姓名', '学号', '学院', '项目名称', '签到时间', '签退时间', '时长(小时)', '确认人'])

    for i, c in enumerate(checkins, 1):
        ws.append([
            i, c.user.real_name, c.user.student_id, c.user.college_name,
            c.project.title if c.project else '',
            c.sign_in_time.strftime('%Y-%m-%d %H:%M') if c.sign_in_time else '',
            c.sign_out_time.strftime('%Y-%m-%d %H:%M') if c.sign_out_time else '',
            c.duration_hours or 0,
            c.checker.real_name if c.checker else '',
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 15

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, '志愿者时长记录.xlsx'
